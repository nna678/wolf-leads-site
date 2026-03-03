#!/usr/bin/env python3
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parent
PACKAGE = BASE / "google_ads_editor_package_2026-02-23_hyper_12000_v2_no_door"
KEY_XLSX = BASE / "key.xlsx"
KEY_NEG_XLSX = BASE / "key_neg1.xlsx"

sys.path.insert(0, str(BASE / ".pydeps"))
import openpyxl  # type: ignore

CAMPAIGN = "SZ_Search_BayArea_Hyper_ExactPhrase_2026"
FINAL_URL = "https://appliance-pros.netlify.app/subzero/"
START_DATE = "2026-02-23"

FORBIDDEN = {"door", "gasket", "seal", "hinge", "handle", "cost", "quote", "pricing", "estimate"}

SPECS_RAW = """
Geo San Francisco Repair|42|130|geo|repair||san francisco
Geo San Jose Repair|42|130|geo|repair||san jose
Geo Oakland Repair|41|120|geo|repair||oakland
Geo Palo Alto Repair|41|115|geo|repair||palo alto
Geo Mountain View Repair|41|115|geo|repair||mountain view
Geo Sunnyvale Repair|41|110|geo|repair||sunnyvale
Geo San Mateo Repair|41|110|geo|repair||san mateo
Geo Redwood City Repair|41|115|geo|repair||redwood city
Geo Santa Clara Repair|41|115|geo|repair||santa clara
Geo Fremont Repair|41|120|geo|repair||fremont
Wine Cooler Repair|37|95|wine|repair|wine cooler|
Wine Cooler Service|36|70|wine|service|wine cooler|
Wine Cooler Not Cooling|39|90|wine|not_cooling|wine cooler|
Wine Cooler Not Working|39|75|wine|not_working|wine cooler|
Ice Maker Repair|41|105|ice|repair|ice maker|
Ice Maker Service|39|80|ice|service|ice maker|
Ice Maker Near Me|42|75|ice|near_me|ice maker|
Ice Maker Technician|40|70|ice|technician|ice maker|
Ice Maker Not Working|42|100|ice|not_working|ice maker|
Ice Maker Not Making Ice|42|95|ice|not_making_ice|ice maker|
Ice Maker Leaking|42|70|ice|leaking|ice maker|
Ice Dispenser Not Working|42|70|ice|ice_dispenser_not_working|ice maker|
Freezer Repair|39|95|freezer|repair|freezer|
Freezer Service|37|75|freezer|service|freezer|
Freezer Near Me|41|70|freezer|near_me|freezer|
Freezer Technician|38|65|freezer|technician|freezer|
Freezer Not Cooling|42|95|freezer|not_cooling|freezer|
Freezer Not Freezing|42|85|freezer|not_freezing|freezer|
Freezer Not Working|42|85|freezer|not_working|freezer|
Freezer Leaking|42|65|freezer|leaking|freezer|
Fridge Repair|39|120|fridge|repair|fridge|
Fridge Service|37|90|fridge|service|fridge|
Fridge Near Me|41|85|fridge|near_me|fridge|
Fridge Technician|38|80|fridge|technician|fridge|
Fridge Not Cooling|42|110|fridge|not_cooling|fridge|
Fridge Not Freezing|42|85|fridge|not_freezing|fridge|
Fridge Not Working|42|90|fridge|not_working|fridge|
Fridge Leaking|42|75|fridge|leaking|fridge|
Fridge Warm|40|65|fridge|warm|fridge|
Refrigerator Repair|40|130|refrigerator|repair|refrigerator|
Refrigerator Service|38|100|refrigerator|service|refrigerator|
Refrigerator Near Me|42|90|refrigerator|near_me|refrigerator|
Refrigerator Same Day|45|80|refrigerator|same_day|refrigerator|
Refrigerator Technician|39|85|refrigerator|technician|refrigerator|
Refrigerator Not Cooling|43|120|refrigerator|not_cooling|refrigerator|
Refrigerator Not Freezing|43|90|refrigerator|not_freezing|refrigerator|
Refrigerator Not Working|43|95|refrigerator|not_working|refrigerator|
Refrigerator Leaking|43|80|refrigerator|leaking|refrigerator|
Refrigerator Warm|41|70|refrigerator|warm|refrigerator|
Refrigerator Error Code|38|70|refrigerator|error_code|refrigerator|
Core Brand Repair|38|110|core|repair||
Core Brand Service|37|90|core|service||
Core Near Me|43|95|core|near_me||
Core Same Day|46|85|core|same_day||
Core Emergency|48|85|core|emergency||
Core Technician|39|90|core|technician||
Core Repairman|39|75|core|repairman||
Core Mechanic|38|70|core|mechanic||
Core Specialist|38|70|core|specialist||
Core Not Cooling|43|80|core|not_cooling||
Core Not Freezing|43|70|core|not_freezing||
Core Not Working|43|80|core|not_working||
Core Leaking|43|70|core|leaking||
Core Warm|40|60|core|warm||
Core Beeping Noise|38|60|core|beeping_noise||
Core Error Code|38|60|core|error_code||
""".strip()

CITIES = [
    "san francisco",
    "san jose",
    "oakland",
    "palo alto",
    "mountain view",
    "sunnyvale",
    "san mateo",
    "redwood city",
    "santa clara",
    "fremont",
]

CITY_GROUP = {
    "san francisco": "Geo San Francisco Repair",
    "san jose": "Geo San Jose Repair",
    "oakland": "Geo Oakland Repair",
    "palo alto": "Geo Palo Alto Repair",
    "mountain view": "Geo Mountain View Repair",
    "sunnyvale": "Geo Sunnyvale Repair",
    "san mateo": "Geo San Mateo Repair",
    "redwood city": "Geo Redwood City Repair",
    "santa clara": "Geo Santa Clara Repair",
    "fremont": "Geo Fremont Repair",
}

INTENTS = {
    "repair": ["repair", "repair service", "service", "appliance repair", "repair company", "repair experts"],
    "service": ["service", "service appointment", "service company", "maintenance service", "service visit"],
    "near_me": ["repair near me", "service near me", "local repair", "local service", "repair in my area"],
    "same_day": ["same day repair", "same day service", "today repair", "today service", "fast same day repair"],
    "emergency": ["emergency repair", "emergency service", "urgent repair", "urgent service", "asap repair"],
    "technician": ["technician", "repair technician", "service technician", "appliance technician", "refrigeration technician"],
    "repairman": ["repairman", "local repairman", "repairman near me", "appliance repairman", "service repairman"],
    "mechanic": ["mechanic", "appliance mechanic", "refrigeration mechanic", "local mechanic", "mechanic near me"],
    "specialist": ["specialist", "repair specialist", "service specialist", "appliance specialist", "refrigeration specialist"],
    "not_cooling": ["not cooling", "stopped cooling", "cooling problem", "not cold", "warming up"],
    "not_freezing": ["not freezing", "freezing problem", "not freezing well", "stopped freezing"],
    "not_working": ["not working", "stopped working", "not turning on", "won t start", "dead unit"],
    "leaking": ["leaking", "water leak", "leak repair", "water dripping", "leaking water"],
    "warm": ["running warm", "too warm", "warm inside", "temperature high", "not cold enough"],
    "beeping_noise": ["beeping", "noise problem", "compressor noise", "fan noise", "buzzing noise"],
    "error_code": ["error code", "fault code", "code error", "diagnostic", "diagnostics"],
    "not_making_ice": ["not making ice", "no ice", "stopped making ice", "not producing ice"],
    "ice_dispenser_not_working": ["ice dispenser not working", "not dispensing ice", "ice dispenser issue", "ice not dispensing"],
}


def specs():
    out = []
    for line in SPECS_RAW.splitlines():
        name, cpc, target, family, intent, appliance, city = line.split("|")
        out.append(
            {
                "name": name,
                "cpc": float(cpc),
                "target": int(target),
                "family": family,
                "intent": intent,
                "appliance": appliance or None,
                "city": city or None,
            }
        )
    return out


def nrm(s: str) -> str:
    s = s.lower().strip().replace("sub-zero", "sub zero").replace("subzero", "sub zero").replace("’", "'")
    s = re.sub(r"[^a-z0-9' ]+", " ", s)
    return " ".join(s.split())


def canon(s: str) -> str:
    s = nrm(s).replace("sub zero", "subzero")
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return " ".join(s.split())


def is_forbidden(s: str) -> bool:
    t = nrm(s)
    return any(x in t for x in FORBIDDEN)


def subjects(appliance):
    if appliance == "refrigerator":
        return ["sub zero refrigerator", "sub zero built in refrigerator", "sub zero fridge"]
    if appliance == "fridge":
        return ["sub zero fridge", "sub zero refrigerator", "sub zero built in fridge"]
    if appliance == "freezer":
        return ["sub zero freezer", "sub zero freezer unit"]
    if appliance == "ice maker":
        return ["sub zero ice maker", "sub zero ice machine", "sub zero ice dispenser"]
    if appliance == "wine cooler":
        return ["sub zero wine cooler", "sub zero wine refrigerator", "sub zero wine fridge"]
    return ["sub zero", "sub zero appliance", "sub zero unit"]


def detect_geo(s):
    t = nrm(s)
    for city in CITIES:
        if city in t:
            return CITY_GROUP[city]
    return None


def detect_appliance(s):
    t = nrm(s)
    if "wine cooler" in t or "wine refrigerator" in t or "wine fridge" in t:
        return "wine cooler"
    if "ice maker" in t or "ice machine" in t or "ice dispenser" in t:
        return "ice maker"
    if "freezer" in t:
        return "freezer"
    if "fridge" in t:
        return "fridge"
    if "refrigerator" in t:
        return "refrigerator"
    return None


def detect_intent(s):
    t = nrm(s)
    if "same day" in t or "today" in t:
        return "same_day"
    if "emergency" in t or "urgent" in t or "asap" in t:
        return "emergency"
    if "near me" in t or "my area" in t or "local" in t:
        return "near_me"
    if "repairman" in t:
        return "repairman"
    if "mechanic" in t:
        return "mechanic"
    if "specialist" in t:
        return "specialist"
    if "technician" in t or "service tech" in t:
        return "technician"
    if "not making ice" in t or "no ice" in t:
        return "not_making_ice"
    if "ice dispenser not working" in t or "not dispensing ice" in t:
        return "ice_dispenser_not_working"
    if "not cooling" in t or "stopped cooling" in t or "not cold" in t:
        return "not_cooling"
    if "not freezing" in t:
        return "not_freezing"
    if "not working" in t or "stopped working" in t or "not turning on" in t:
        return "not_working"
    if "leak" in t or "leaking" in t:
        return "leaking"
    if "warm" in t or "temperature" in t:
        return "warm"
    if "beeping" in t or "noise" in t or "compressor" in t or "fan" in t:
        return "beeping_noise"
    if "error code" in t or "fault code" in t or "diagnostic" in t:
        return "error_code"
    if "service" in t:
        return "service"
    return "repair"


def resolve_group(s):
    geo = detect_geo(s)
    if geo:
        return geo
    a = detect_appliance(s)
    i = detect_intent(s)
    map_ = {
        ("refrigerator", "repair"): "Refrigerator Repair",
        ("refrigerator", "service"): "Refrigerator Service",
        ("refrigerator", "near_me"): "Refrigerator Near Me",
        ("refrigerator", "same_day"): "Refrigerator Same Day",
        ("refrigerator", "emergency"): "Refrigerator Same Day",
        ("refrigerator", "technician"): "Refrigerator Technician",
        ("refrigerator", "repairman"): "Refrigerator Technician",
        ("refrigerator", "mechanic"): "Refrigerator Technician",
        ("refrigerator", "specialist"): "Refrigerator Technician",
        ("refrigerator", "not_cooling"): "Refrigerator Not Cooling",
        ("refrigerator", "not_freezing"): "Refrigerator Not Freezing",
        ("refrigerator", "not_working"): "Refrigerator Not Working",
        ("refrigerator", "leaking"): "Refrigerator Leaking",
        ("refrigerator", "warm"): "Refrigerator Warm",
        ("refrigerator", "error_code"): "Refrigerator Error Code",
        ("fridge", "repair"): "Fridge Repair",
        ("fridge", "service"): "Fridge Service",
        ("fridge", "near_me"): "Fridge Near Me",
        ("fridge", "technician"): "Fridge Technician",
        ("fridge", "repairman"): "Fridge Technician",
        ("fridge", "mechanic"): "Fridge Technician",
        ("fridge", "specialist"): "Fridge Technician",
        ("fridge", "not_cooling"): "Fridge Not Cooling",
        ("fridge", "not_freezing"): "Fridge Not Freezing",
        ("fridge", "not_working"): "Fridge Not Working",
        ("fridge", "leaking"): "Fridge Leaking",
        ("fridge", "warm"): "Fridge Warm",
        ("freezer", "repair"): "Freezer Repair",
        ("freezer", "service"): "Freezer Service",
        ("freezer", "near_me"): "Freezer Near Me",
        ("freezer", "technician"): "Freezer Technician",
        ("freezer", "repairman"): "Freezer Technician",
        ("freezer", "mechanic"): "Freezer Technician",
        ("freezer", "specialist"): "Freezer Technician",
        ("freezer", "not_cooling"): "Freezer Not Cooling",
        ("freezer", "not_freezing"): "Freezer Not Freezing",
        ("freezer", "not_working"): "Freezer Not Working",
        ("freezer", "leaking"): "Freezer Leaking",
        ("ice maker", "repair"): "Ice Maker Repair",
        ("ice maker", "service"): "Ice Maker Service",
        ("ice maker", "near_me"): "Ice Maker Near Me",
        ("ice maker", "technician"): "Ice Maker Technician",
        ("ice maker", "repairman"): "Ice Maker Technician",
        ("ice maker", "mechanic"): "Ice Maker Technician",
        ("ice maker", "specialist"): "Ice Maker Technician",
        ("ice maker", "not_working"): "Ice Maker Not Working",
        ("ice maker", "not_making_ice"): "Ice Maker Not Making Ice",
        ("ice maker", "ice_dispenser_not_working"): "Ice Dispenser Not Working",
        ("ice maker", "leaking"): "Ice Maker Leaking",
        ("ice maker", "not_cooling"): "Ice Maker Not Working",
        ("wine cooler", "repair"): "Wine Cooler Repair",
        ("wine cooler", "service"): "Wine Cooler Service",
        ("wine cooler", "not_cooling"): "Wine Cooler Not Cooling",
        ("wine cooler", "not_working"): "Wine Cooler Not Working",
        (None, "repair"): "Core Brand Repair",
        (None, "service"): "Core Brand Service",
        (None, "near_me"): "Core Near Me",
        (None, "same_day"): "Core Same Day",
        (None, "emergency"): "Core Emergency",
        (None, "technician"): "Core Technician",
        (None, "repairman"): "Core Repairman",
        (None, "mechanic"): "Core Mechanic",
        (None, "specialist"): "Core Specialist",
        (None, "not_cooling"): "Core Not Cooling",
        (None, "not_freezing"): "Core Not Freezing",
        (None, "not_working"): "Core Not Working",
        (None, "leaking"): "Core Leaking",
        (None, "warm"): "Core Warm",
        (None, "beeping_noise"): "Core Beeping Noise",
        (None, "error_code"): "Core Error Code",
    }
    return map_.get((a, i), "Core Brand Repair")


def add(pool, seen, group, phrase):
    phrase = nrm(phrase)
    if not phrase or is_forbidden(phrase) or "sub zero" not in phrase:
        return
    c = canon(phrase)
    if c in seen[group]:
        return
    seen[group].add(c)
    pool[group].append(phrase)


def fallback(appliance, intent, city, idx):
    sub = subjects(appliance)[idx % len(subjects(appliance))]
    if city:
        seq = [
            f"{sub} repair {city}",
            f"{city} {sub} service",
            f"{sub} technician {city}",
            f"same day {sub} repair {city}",
            f"{sub} not cooling repair {city}",
        ]
        return seq[idx % len(seq)]
    if intent in {"not_cooling", "not_freezing", "not_working", "leaking", "warm", "beeping_noise", "error_code", "not_making_ice", "ice_dispenser_not_working"}:
        sym = INTENTS[intent][idx % len(INTENTS[intent])]
        tail = ["repair", "service", "technician", "same day repair", "emergency service", "diagnostic", "help"]
        mod = ["near me", "local", "bay area", "today", "urgent", "fast response"]
        if idx % 3 == 0:
            return f"{sub} {sym} {tail[(idx // 3) % len(tail)]} {mod[(idx // 7) % len(mod)]}"
        if idx % 3 == 1:
            return f"{tail[(idx // 2) % len(tail)]} for {sub} {sym} {mod[(idx // 5) % len(mod)]}"
        return f"{sub} {sym} issue {mod[(idx // 4) % len(mod)]}"
    if intent == "near_me":
        m1 = ["near me", "in my area", "local", "nearby", "around me", "close by"]
        m2 = ["repair", "service", "technician", "repair help", "service support", "urgent help"]
        m3 = ["today", "same day", "fast", "priority", "trusted", "professional", "quick response", "available now"]
        return f"{sub} {m2[idx % len(m2)]} {m1[(idx // len(m2)) % len(m1)]} {m3[(idx // (len(m2)*len(m1))) % len(m3)]}"
    if intent == "same_day":
        mods = ["same day", "today", "priority", "fast", "urgent", "immediate"]
        tail = ["repair", "service", "technician visit", "service call", "repair booking", "diagnostic"]
        end = ["near me", "bay area", "local", "available now", "with fast response"]
        return f"{mods[idx % len(mods)]} {sub} {tail[(idx // len(mods)) % len(tail)]} {end[(idx // (len(mods)*len(tail))) % len(end)]}"
    if intent == "emergency":
        mods = ["emergency", "urgent", "asap", "priority", "critical", "rapid"]
        tail = ["repair", "service", "technician", "service call", "repair booking", "diagnostic"]
        end = ["near me", "bay area", "local", "available now", "today"]
        return f"{mods[idx % len(mods)]} {sub} {tail[(idx // len(mods)) % len(tail)]} {end[(idx // (len(mods)*len(tail))) % len(end)]}"
    if intent in {"technician", "repairman", "mechanic", "specialist"}:
        m1 = ["near me", "local", "bay area", "today", "same day", "emergency", "urgent", "in my area"]
        m2 = ["visit", "support", "service", "call", "appointment", "help"]
        if idx % 2:
            return f"{sub} {intent} {m1[idx % len(m1)]} {m2[(idx // len(m1)) % len(m2)]}"
        return f"{intent} for {sub} {m1[idx % len(m1)]} {m2[(idx // len(m1)) % len(m2)]}"
    if intent == "service":
        m1 = ["near me", "local", "bay area", "today", "same day", "urgent", "in my area", "appointment"]
        m2 = ["fast", "trusted", "professional", "local team", "priority", "available now"]
        m3 = ["support", "booking", "visit", "diagnostic", "response", "scheduling", "team"]
        return f"{sub} service {m1[idx % len(m1)]} {m2[(idx // len(m1)) % len(m2)]} {m3[(idx // (len(m1)*len(m2))) % len(m3)]}"
    if intent == "repair":
        m1 = ["near me", "local", "bay area", "today", "same day", "urgent", "in my area", "fast", "trusted"]
        m2 = ["service", "support", "help", "team", "technician", "experts", "company"]
        m3 = ["now", "available", "response", "booking", "visit", "priority", "dispatch"]
        return f"{sub} repair {m1[idx % len(m1)]} {m2[(idx // len(m1)) % len(m2)]} {m3[(idx // (len(m1)*len(m2))) % len(m3)]}"
    a = ["repair", "service", "repair service", "service appointment", "fast repair", "local service"]
    b = ["near me", "bay area", "today", "same day", "urgent", "trusted", "local", "in my area"]
    return f"{sub} {a[idx % len(a)]} {b[(idx // len(a)) % len(b)]}"


def build():
    specs_list = specs()
    total_target = sum(x["target"] for x in specs_list)
    target_exact = 6000
    if total_target != target_exact:
        # Rebalance targets to hit exact volume requirement.
        grow_order = [
            "Core Brand Repair",
            "Core Near Me",
            "Refrigerator Repair",
            "Fridge Repair",
            "Freezer Repair",
            "Ice Maker Repair",
            "Geo San Jose Repair",
            "Geo San Francisco Repair",
        ]
        by_name = {x["name"]: x for x in specs_list}
        if total_target < target_exact:
            need = target_exact - total_target
            i = 0
            while need > 0:
                key = grow_order[i % len(grow_order)]
                by_name[key]["target"] += 1
                need -= 1
                i += 1
        else:
            need = total_target - target_exact
            shrink_order = list(reversed(grow_order))
            i = 0
            while need > 0:
                key = shrink_order[i % len(shrink_order)]
                if by_name[key]["target"] > 10:
                    by_name[key]["target"] -= 1
                    need -= 1
                i += 1
        total_target = sum(x["target"] for x in specs_list)

    pool = {x["name"]: [] for x in specs_list}
    seen = {x["name"]: set() for x in specs_list}

    wb = openpyxl.load_workbook(KEY_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() for h in rows[0]]
    kw_i = next(i for i, h in enumerate(header) if "Keyword" in h)
    seed_seen = set()
    for r in rows[1:]:
        kw = nrm(str(r[kw_i] or ""))
        if not kw:
            continue
        if "sub zero" not in kw and "subzero" not in kw and "sub-zero" not in kw:
            continue
        if is_forbidden(kw):
            continue
        c = canon(kw)
        if c in seed_seen:
            continue
        seed_seen.add(c)
        add(pool, seen, resolve_group(kw), kw)

    for s in specs_list:
        name, fam, intent, appliance, city = s["name"], s["family"], s["intent"], s["appliance"], s["city"]
        if fam == "geo":
            apps = ["refrigerator", "fridge", "freezer", "ice maker", "wine cooler", "appliance"]
            svcs = ["repair", "service", "repair service", "technician", "same day repair", "emergency repair", "not cooling repair", "not freezing repair", "not working repair", "leaking repair"]
            for ap in apps:
                for sv in svcs:
                    add(pool, seen, name, f"sub zero {ap} {sv} {city}")
                    add(pool, seen, name, f"{city} sub zero {ap} {sv}")
                    add(pool, seen, name, f"sub zero {ap} {sv} in {city}")
        else:
            for sub in subjects(appliance):
                for it in INTENTS.get(intent, ["repair"]):
                    add(pool, seen, name, f"{sub} {it}")
                    add(pool, seen, name, f"{it} {sub}")
                for mod in ["near me", "local", "bay area", "same day", "emergency", "today"]:
                    if intent in {"near_me", "same_day", "emergency"}:
                        add(pool, seen, name, f"{sub} {INTENTS[intent][0]}")
                    else:
                        add(pool, seen, name, f"{sub} repair {mod}")
                        add(pool, seen, name, f"{sub} service {mod}")

    selected = {x["name"]: [] for x in specs_list}
    for s in specs_list:
        name, target = s["name"], s["target"]
        local_seen = set()
        for kw in pool[name]:
            c = canon(kw)
            if c in local_seen:
                continue
            local_seen.add(c)
            selected[name].append(kw)
            if len(selected[name]) >= target:
                break
        i = 0
        while len(selected[name]) < target:
            kw = nrm(fallback(s["appliance"], s["intent"], s["city"], i))
            c = canon(kw)
            if c not in local_seen and not is_forbidden(kw):
                local_seen.add(c)
                selected[name].append(kw)
            i += 1
            if i > 30000:
                raise RuntimeError(f"Cannot fill {name}")

    exact = [(s["name"], kw) for s in specs_list for kw in selected[s["name"]]]
    if len(exact) != total_target:
        raise RuntimeError(f"Expected {total_target}, got {len(exact)}")

    PACKAGE.mkdir(parents=True, exist_ok=True)
    with (PACKAGE / "01_campaign.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Campaign", "Campaign type", "Campaign daily budget", "Bid strategy type", "Networks", "Languages", "Start date", "Campaign status", "Comment"])
        w.writerow([CAMPAIGN, "Search", "900.00", "Manual CPC", "Google Search;Search Partners", "en", START_DATE, "Enabled", "Hyper-segmented launch: 6000 exact + 6000 phrase; no broad; no door; no cost/quote intent."])

    with (PACKAGE / "02_ad_groups.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Campaign", "Ad Group", "Max CPC", "Campaign Status", "Ad Group Status", "Comment"])
        for s in specs_list:
            w.writerow([CAMPAIGN, s["name"], f"{s['cpc']:.2f}", "Enabled", "Enabled", f"Hyper-segmentation: {s['family']} | {s['intent']}"])

    cpc = {s["name"]: f"{s['cpc']:.2f}" for s in specs_list}
    kw_header = ["Campaign", "Ad Group", "Keyword", "Criterion Type", "Max CPC", "Final URL", "Campaign Status", "Ad Group Status", "Status", "Comment"]
    kw_rows = []
    for ag, kw in exact:
        for ct in ("Exact", "Phrase"):
            kw_rows.append([CAMPAIGN, ag, kw, ct, cpc[ag], FINAL_URL, "Enabled", "Enabled", "Enabled", f"{ag} | {ct}"])
    for fn in ("03_keywords_12000_master.csv", "03a_keywords_exact_phrase.csv"):
        with (PACKAGE / fn).open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(kw_header)
            w.writerows(kw_rows)

    print("done", PACKAGE)
    print("ad_groups", len(specs_list), "exact", len(exact), "rows", len(kw_rows))


if __name__ == "__main__":
    build()
