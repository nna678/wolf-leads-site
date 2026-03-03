#!/usr/bin/env python3
"""
Build Google Ads package: 1 campaign, Exact + Phrase only, 5000 unique keywords = 10000 rows.
Uses key.xlsx for seed keywords and maps to 51 themed ad groups for max Quality Score.
"""
from pathlib import Path
import csv
import re

BASE = Path(__file__).resolve().parent
PACKAGE = BASE / "google_ads_editor_package_2026-02-23_deep_5000"
KEY_XLSX = BASE / "key.xlsx"

CAMPAIGN_NAME = "SZ_Search_BayArea_ExactPhrase_2026"
FINAL_URL = "https://appliancepros.netlify.app/subzero/"

# Our 51 ad groups (Core only) with default Max CPC
AD_GROUPS_CORE = [
    ("Core Brand Repair", 36.00),
    ("Core Brand Service", 36.00),
    ("Core Cost Quote", 34.00),
    ("Core Emergency", 48.00),
    ("Core Near Me", 44.00),
    ("Door Seal Repair", 36.00),
    ("Freezer Repair", 37.00),
    ("Geo Alameda Repair", 39.00),
    ("Geo Belmont Repair", 39.00),
    ("Geo Berkeley Repair", 39.00),
    ("Geo Burlingame Repair", 39.00),
    ("Geo Campbell Repair", 39.00),
    ("Geo CastroValley Repair", 39.00),
    ("Geo Concord Repair", 39.00),
    ("Geo Cupertino Repair", 39.00),
    ("Geo DalyCity Repair", 39.00),
    ("Geo Emeryville Repair", 39.00),
    ("Geo FosterCity Repair", 39.00),
    ("Geo Fremont Repair", 39.00),
    ("Geo Hayward Repair", 39.00),
    ("Geo Livermore Repair", 39.00),
    ("Geo LosAltos Repair", 39.00),
    ("Geo LosGatos Repair", 39.00),
    ("Geo MenloPark Repair", 39.00),
    ("Geo Millbrae Repair", 39.00),
    ("Geo Milpitas Repair", 39.00),
    ("Geo MountainView Repair", 39.00),
    ("Geo Novato Repair", 39.00),
    ("Geo Oakland Repair", 39.00),
    ("Geo PaloAlto Repair", 39.00),
    ("Geo Pleasanton Repair", 39.00),
    ("Geo RedwoodCity Repair", 39.00),
    ("Geo Richmond Repair", 39.00),
    ("Geo SanBruno Repair", 39.00),
    ("Geo SanFrancisco Repair", 39.00),
    ("Geo SanJose Repair", 39.00),
    ("Geo SanLeandro Repair", 39.00),
    ("Geo SanMateo Repair", 39.00),
    ("Geo SanRafael Repair", 39.00),
    ("Geo SantaClara Repair", 39.00),
    ("Geo Sunnyvale Repair", 39.00),
    ("Geo UnionCity Repair", 39.00),
    ("Geo WalnutCreek Repair", 39.00),
    ("Ice Maker Repair", 40.00),
    ("Noise Compressor Fan", 36.00),
    ("Not Cooling Issues", 42.00),
    ("Not Freezing Issues", 40.00),
    ("Not Working Issues", 38.00),
    ("Technician Request", 37.00),
    ("Temperature Issues", 37.00),
    ("Water Leak Repair", 42.00),
    ("Wine Cooler Repair", 35.00),
]

# Geo: key phrase in keyword -> our ad group name
GEO_MAP = {
    "alameda": "Geo Alameda Repair",
    "belmont": "Geo Belmont Repair",
    "berkeley": "Geo Berkeley Repair",
    "burlingame": "Geo Burlingame Repair",
    "campbell": "Geo Campbell Repair",
    "castro valley": "Geo CastroValley Repair",
    "concord": "Geo Concord Repair",
    "cupertino": "Geo Cupertino Repair",
    "daly city": "Geo DalyCity Repair",
    "emeryville": "Geo Emeryville Repair",
    "foster city": "Geo FosterCity Repair",
    "fremont": "Geo Fremont Repair",
    "hayward": "Geo Hayward Repair",
    "livermore": "Geo Livermore Repair",
    "los altos": "Geo LosAltos Repair",
    "los gatos": "Geo LosGatos Repair",
    "menlo park": "Geo MenloPark Repair",
    "millbrae": "Geo Millbrae Repair",
    "milpitas": "Geo Milpitas Repair",
    "mountain view": "Geo MountainView Repair",
    "novato": "Geo Novato Repair",
    "oakland": "Geo Oakland Repair",
    "palo alto": "Geo PaloAlto Repair",
    "pleasanton": "Geo Pleasanton Repair",
    "redwood city": "Geo RedwoodCity Repair",
    "richmond": "Geo Richmond Repair",
    "san bruno": "Geo SanBruno Repair",
    "san francisco": "Geo SanFrancisco Repair",
    "san jose": "Geo SanJose Repair",
    "san leandro": "Geo SanLeandro Repair",
    "san mateo": "Geo SanMateo Repair",
    "san rafael": "Geo SanRafael Repair",
    "santa clara": "Geo SantaClara Repair",
    "sunnyvale": "Geo Sunnyvale Repair",
    "union city": "Geo UnionCity Repair",
    "walnut creek": "Geo WalnutCreek Repair",
}

# key.xlsx ad_group (lower) -> our ad group
KEY_AG_TO_OUR_AG = {
    "refrigerator repair near me": "Core Near Me",
    "refrigerator repair service near me": "Core Near Me",
    "refrigerator repair service": "Core Brand Repair",
    "refrigerator repair": "Core Brand Repair",
    "refrigerator fix": "Core Brand Repair",
    "refrigerator repairman": "Technician Request",
    "refrigerator repairman near me": "Core Near Me",
    "refrigerator mechanic near me": "Core Near Me",
    "refrigerator technician": "Technician Request",
    "appliance repair": "Core Brand Repair",
    "appliance repair near me": "Core Near Me",
    "appliance repair service": "Core Brand Repair",
    "appliance repair service near me": "Core Near Me",
    "appliance repair my area": "Core Near Me",
    "appliance repairs": "Core Brand Repair",
    "appliance service": "Core Brand Service",
    "appliance service near me": "Core Near Me",
    "appliance service center": "Core Brand Service",
    "repair near me": "Core Near Me",
    "repair service near me": "Core Near Me",
    "repair service": "Core Brand Repair",
    "repair service center": "Core Brand Repair",
    "repair center near me": "Core Near Me",
    "repair center": "Core Brand Repair",
    "repair my area": "Core Near Me",
    "service near me": "Core Near Me",
    "service center near me": "Core Near Me",
    "service center": "Core Brand Service",
    "service my area": "Core Near Me",
    "repairman near me": "Core Near Me",
    "repairman": "Technician Request",
    "technician near me": "Core Near Me",
    "technician": "Technician Request",
    "service technician": "Technician Request",
    "service technician near me": "Core Near Me",
    "repairs near me": "Core Near Me",
    "services near me": "Core Near Me",
    "local appliance repair": "Core Near Me",
    "local appliance repair near me": "Core Near Me",
    "local repair service": "Core Near Me",
    "repair san jose": "Geo SanJose Repair",
    "repair palo alto": "Geo PaloAlto Repair",
    "repair san mateo": "Geo SanMateo Repair",
    "repair SF": "Geo SanFrancisco Repair",
    "service san jose": "Geo SanJose Repair",
    "service palo alto": "Geo PaloAlto Repair",
    "service san mateo": "Geo SanMateo Repair",
    "service SF": "Geo SanFrancisco Repair",
    # Freezer
    "freezer repair": "Freezer Repair",
    "freezer repair near me": "Core Near Me",
    "freezer repair service": "Freezer Repair",
    "freezer repairs": "Freezer Repair",
    "freezer repairs near me": "Core Near Me",
    "freezer service": "Freezer Repair",
    "freezer service near me": "Core Near Me",
    "freezer fix": "Freezer Repair",
    "freezer not cooling": "Not Cooling Issues",
    "freezer not freezing": "Not Freezing Issues",
    "freezer not working": "Not Working Issues",
    "freezer not cold": "Not Cooling Issues",
    "freezer not freezing 2": "Not Freezing Issues",
    "freezer not getting cold": "Not Cooling Issues",
    "freezer not making ice": "Ice Maker Repair",
    "freezer stopped working": "Not Working Issues",
    "freezer problem": "Not Working Issues",
    "freezer problems": "Not Working Issues",
    "freezer leaking": "Water Leak Repair",
    "freezer ice maker not working": "Ice Maker Repair",
    "freezer ice maker not making ice": "Ice Maker Repair",
    "freezer ice maker not dispensing": "Ice Maker Repair",
    "freezer ice maker repair": "Ice Maker Repair",
    "freezer water dispenser not working": "Water Leak Repair",
    "freezer door not sealing": "Door Seal Repair",
    "freezer door broken": "Door Seal Repair",
    "freezer door stuck": "Door Seal Repair",
    "freezer handle loose": "Door Seal Repair",
    "freezer light not working": "Temperature Issues",
    "freezer not defrosting": "Not Freezing Issues",
    "freezer ice build up": "Ice Maker Repair",
    "freezer beeping": "Noise Compressor Fan",
    "freezer too cold": "Temperature Issues",
    "freezer stuck": "Not Working Issues",
    "freezer stopped cooling": "Not Cooling Issues",
    # Fridge
    "fridge repair": "Core Brand Repair",
    "fridge repair near me": "Core Near Me",
    "fridge repair service": "Core Brand Repair",
    "fridge repair service near me": "Core Near Me",
    "fridge repairs": "Core Brand Repair",
    "fridge repairs near me": "Core Near Me",
    "fridge service": "Core Brand Service",
    "fridge service near me": "Core Near Me",
    "fridge fix": "Core Brand Repair",
    "fridge not cooling": "Not Cooling Issues",
    "fridge not cold": "Not Cooling Issues",
    "fridge not freezing": "Not Freezing Issues",
    "fridge not working": "Not Working Issues",
    "fridge not making ice": "Ice Maker Repair",
    "fridge ice maker not working": "Ice Maker Repair",
    "fridge ice maker repair": "Ice Maker Repair",
    "fridge ice maker not making ice": "Ice Maker Repair",
    "fridge ice maker not dispensing": "Ice Maker Repair",
    "fridge leaking": "Water Leak Repair",
    "fridge water dispenser not working": "Water Leak Repair",
    "fridge water leak": "Water Leak Repair",
    "fridge light not working": "Temperature Issues",
    "fridge door": "Door Seal Repair",
    "fridge handle loose": "Door Seal Repair",
    "fridge not defrosting": "Not Freezing Issues",
    "fridge stopped cooling": "Not Cooling Issues",
    "fridge stopped working": "Not Working Issues",
    "fridge technician": "Technician Request",
    "fridge repairman": "Technician Request",
    "fridge too cold": "Temperature Issues",
    "fridge warm": "Temperature Issues",
    "fridge noise": "Noise Compressor Fan",
    "fridge beeping": "Noise Compressor Fan",
    "fridge maintenance": "Core Brand Service",
    "fridge drain blocked": "Water Leak Repair",
    "fridge ice maker leaking": "Water Leak Repair",
    "fridge not sealing": "Door Seal Repair",
    "fridge not turning on": "Not Working Issues",
    "fridge not dispensing": "Water Leak Repair",
    "fridge freezer not cooling": "Not Cooling Issues",
    "fridge freezer not fridge": "Temperature Issues",
    "fridge freezing up": "Not Freezing Issues",
    "fridge wont make ice": "Ice Maker Repair",
    "fridge not getting cold": "Not Cooling Issues",
    "fridge not making ice 2": "Ice Maker Repair",
    "fridge ice maker problems": "Ice Maker Repair",
    "fridge ice maker slow": "Ice Maker Repair",
    "fridge ice maker stopped working": "Ice Maker Repair",
    "fridge ice not coming": "Ice Maker Repair",
    "fridge leaking 2": "Water Leak Repair",
    "fridge not cooling 2": "Not Cooling Issues",
    "fridge not cooling but light": "Not Cooling Issues",
    "fridge not freezing": "Not Freezing Issues",
    "fridge not fridge": "Temperature Issues",
    "fridge problem": "Not Working Issues",
    "fridge problems": "Not Working Issues",
    "fridge broken": "Not Working Issues",
    "fridge fix 2": "Core Brand Repair",
    "fridge isn t making": "Ice Maker Repair",
    "fridge stopped making": "Ice Maker Repair",
    # Refrigerator (long tail)
    "refrigerator not cooling": "Not Cooling Issues",
    "refrigerator not cold": "Not Cooling Issues",
    "refrigerator not freezing": "Not Freezing Issues",
    "refrigerator not working": "Not Working Issues",
    "refrigerator not making ice": "Ice Maker Repair",
    "refrigerator ice maker not working": "Ice Maker Repair",
    "refrigerator ice maker repair": "Ice Maker Repair",
    "refrigerator ice maker repair near me": "Core Near Me",
    "refrigerator leaking": "Water Leak Repair",
    "refrigerator water dispenser not working": "Water Leak Repair",
    "refrigerator water leak": "Water Leak Repair",
    "refrigerator door not closing": "Door Seal Repair",
    "refrigerator door repair": "Door Seal Repair",
    "refrigerator door not sealing": "Door Seal Repair",
    "refrigerator handle loose": "Door Seal Repair",
    "refrigerator freezer not cooling": "Not Cooling Issues",
    "refrigerator freezer not freezing": "Not Freezing Issues",
    "refrigerator freezer not working": "Not Working Issues",
    "refrigerator freezer repair": "Freezer Repair",
    "refrigerator freezer ice maker not working": "Ice Maker Repair",
    "refrigerator stopped cooling": "Not Cooling Issues",
    "refrigerator stopped working": "Not Working Issues",
    "refrigerator repair 2": "Core Brand Repair",
    "refrigerator repair my area": "Core Near Me",
    "refrigerator repair service near me": "Core Near Me",
    "refrigerator service": "Core Brand Service",
    "refrigerator service near me": "Core Near Me",
    "refrigerator services near me": "Core Near Me",
    "refrigerator servicing": "Core Brand Service",
    "refrigerator maintenance": "Core Brand Service",
    "refrigerator temperature": "Temperature Issues",
    "refrigerator too cold": "Temperature Issues",
    "refrigerator warm": "Temperature Issues",
    "refrigerator beeping": "Noise Compressor Fan",
    "refrigerator noise": "Noise Compressor Fan",
    "refrigerator compressor not running": "Not Working Issues",
    "refrigerator does not make ice": "Ice Maker Repair",
    "refrigerator does not cool": "Not Cooling Issues",
    "refrigerator doesn t cool": "Not Cooling Issues",
    "refrigerator not blowing cold air": "Not Cooling Issues",
    "refrigerator not defrosting": "Not Freezing Issues",
    "refrigerator not dispensing": "Water Leak Repair",
    "refrigerator not getting cold": "Not Cooling Issues",
    "refrigerator not making ice 2": "Ice Maker Repair",
    "refrigerator not running": "Not Working Issues",
    "refrigerator dripping water": "Water Leak Repair",
    "refrigerator water dispenser leaking": "Water Leak Repair",
    "refrigerator water dispenser quit working": "Water Leak Repair",
    "refrigerator water not coming": "Water Leak Repair",
    "refrigerator won t cool": "Not Cooling Issues",
    "refrigerator wont cool": "Not Cooling Issues",
    "refrigerator wont get cold": "Not Cooling Issues",
    "refrigerator no ice": "Ice Maker Repair",
    "refrigerator no power": "Not Working Issues",
    "refrigerator light not working": "Temperature Issues",
    "refrigerator light problems": "Temperature Issues",
    "refrigerator light wont come on": "Temperature Issues",
    "refrigerator ice dispenser not working": "Ice Maker Repair",
    "refrigerator ice maker not making ice": "Ice Maker Repair",
    "refrigerator ice maker not dispensing": "Ice Maker Repair",
    "refrigerator ice maker not working": "Ice Maker Repair",
    "refrigerator ice maker problems": "Ice Maker Repair",
    "refrigerator ice maker repair near me": "Core Near Me",
    "refrigerator ice maker fix": "Ice Maker Repair",
    "refrigerator ice maker leaking": "Water Leak Repair",
    "refrigerator ice maker slow": "Ice Maker Repair",
    "refrigerator ice maker stopped working": "Ice Maker Repair",
    "refrigerator ice maker stopped making": "Ice Maker Repair",
    "refrigerator ice maker no water": "Water Leak Repair",
    "refrigerator ice maker not getting water": "Water Leak Repair",
    "refrigerator ice maker not dumping ice": "Ice Maker Repair",
    "refrigerator ice maker quit working": "Ice Maker Repair",
    "refrigerator ice maker freezing up": "Ice Maker Repair",
    "refrigerator ice build up": "Ice Maker Repair",
    "refrigerator freezer door broken": "Door Seal Repair",
    "refrigerator freezer handle loose": "Door Seal Repair",
    "refrigerator freezer ice build up": "Ice Maker Repair",
    "refrigerator freezer ice maker problems": "Ice Maker Repair",
    "refrigerator freezer light not working": "Temperature Issues",
    "refrigerator freezer not cooling 2": "Not Cooling Issues",
    "refrigerator freezer not working 2": "Not Working Issues",
    "refrigerator freezer problems": "Not Working Issues",
    "refrigerator freezer warm": "Temperature Issues",
    "refrigerator freezer not freezing": "Not Freezing Issues",
    "refrigerator door not cooling": "Door Seal Repair",
    "refrigerator door not working": "Door Seal Repair",
    "refrigerator door problems": "Door Seal Repair",
    "refrigerator door service": "Door Seal Repair",
    "refrigerator freezing up": "Not Freezing Issues",
    "refrigerator ice dispenser door repair": "Door Seal Repair",
    "refrigerator ice machine not working": "Ice Maker Repair",
    "refrigerator ice maker broken": "Ice Maker Repair",
    "refrigerator ice maker doesn t work": "Ice Maker Repair",
    "refrigerator ice maker not dumping ice": "Ice Maker Repair",
    "refrigerator icemaker not working": "Ice Maker Repair",
    "refrigerator issues": "Not Working Issues",
    "refrigerator led lights": "Temperature Issues",
    "refrigerator cooling off": "Not Cooling Issues",
    "refrigerator ajar light stays on": "Door Seal Repair",
    "refrigerator runs continuously": "Noise Compressor Fan",
    "refrigerator stop making ice": "Ice Maker Repair",
    "refrigerator stopped making": "Ice Maker Repair",
    "refrigerator not getting cold 2": "Not Cooling Issues",
    "refrigerator not cooling 2": "Not Cooling Issues",
    # Ice maker
    "ice maker repair": "Ice Maker Repair",
    "ice maker repair 2": "Ice Maker Repair",
    "ice maker repair near me": "Core Near Me",
    "ice maker repair service": "Ice Maker Repair",
    "ice maker not working": "Ice Maker Repair",
    "ice maker not making ice": "Ice Maker Repair",
    "ice maker fix": "Ice Maker Repair",
    "ice maker fix 2": "Ice Maker Repair",
    "ice maker leaking": "Water Leak Repair",
    "ice maker not dispensing": "Ice Maker Repair",
    "ice maker not dumping": "Ice Maker Repair",
    "ice maker not filling": "Ice Maker Repair",
    "ice maker not getting cold": "Ice Maker Repair",
    "ice maker not getting water": "Water Leak Repair",
    "ice maker not making enough": "Ice Maker Repair",
    "ice maker problem": "Ice Maker Repair",
    "ice maker problems": "Ice Maker Repair",
    "ice maker quit working": "Ice Maker Repair",
    "ice maker service": "Ice Maker Repair",
    "ice maker slow": "Ice Maker Repair",
    "ice maker stopped making": "Ice Maker Repair",
    "ice maker stopped working": "Ice Maker Repair",
    "ice maker freezing up": "Ice Maker Repair",
    "ice maker no water": "Water Leak Repair",
    "ice maker water dispenser not working": "Water Leak Repair",
    "ice maker troubleshoot": "Ice Maker Repair",
    "ice maker pouring out": "Ice Maker Repair",
    "ice maker much ice": "Ice Maker Repair",
    "ice maker not crushing": "Ice Maker Repair",
    "ice dispenser not working": "Ice Maker Repair",
    "ice dispenser problem": "Ice Maker Repair",
    "ice dispenser sticking": "Ice Maker Repair",
    "ice drawer problems": "Ice Maker Repair",
    "ice machine not working": "Ice Maker Repair",
    "ice machine not making ice": "Ice Maker Repair",
    "ice machine repair": "Ice Maker Repair",
    "ice cube maker not working": "Ice Maker Repair",
    "ice crusher not working": "Ice Maker Repair",
    "ice flap repair": "Door Seal Repair",
    # Water / leak / door / seal
    "water dispenser not working": "Water Leak Repair",
    "water dispenser not cooling": "Water Leak Repair",
    "water dispenser problems": "Water Leak Repair",
    "water dispenser ice dispenser not working": "Ice Maker Repair",
    "door handle repair": "Door Seal Repair",
    "seal problem": "Door Seal Repair",
    "leaking": "Water Leak Repair",
    "light not working": "Temperature Issues",
    # Generic / maintenance
    "maintenance": "Core Brand Service",
    "maintenance near me": "Core Near Me",
    "maintenance services": "Core Brand Service",
    "care": "Core Brand Service",
    "centric repair": "Core Brand Repair",
    "best": "Core Brand Repair",
    "fix": "Core Brand Repair",
    "repair": "Core Brand Repair",
    "repair 1": "Core Brand Repair",
    "repair 2": "Core Brand Repair",
    "repairs": "Core Brand Repair",
    "repairs 2": "Core Brand Repair",
    "service": "Core Brand Service",
    "service 2": "Core Brand Service",
    "services": "Core Brand Service",
    "services 2": "Core Brand Service",
    "servicing": "Core Brand Service",
    "not cold": "Not Cooling Issues",
    "not cooling": "Not Cooling Issues",
    "not freezing": "Not Freezing Issues",
    "not getting cold": "Not Cooling Issues",
    "not making ice": "Ice Maker Repair",
    "not making ice 1": "Ice Maker Repair",
    "not making ice 2": "Ice Maker Repair",
    "not working": "Not Working Issues",
    "problem": "Not Working Issues",
    "problem 2": "Not Working Issues",
    "problems": "Not Working Issues",
    "stopped cooling": "Not Cooling Issues",
    "stopped making": "Ice Maker Repair",
    "refrigeration repair": "Core Brand Repair",
    "refrigeration repairs": "Core Brand Repair",
    "gas refrigerator door repair": "Door Seal Repair",
    "gate opener repair": "Core Brand Repair",
    "gate repair": "Core Brand Repair",
}


def normalize_kw(s):
    if not s or not isinstance(s, str):
        return ""
    return " ".join(s.lower().split()).strip()


def map_key_ag_to_our_ag(key_ag: str, keyword_lower: str) -> str:
    """Resolve our ad group from key.xlsx ad_group and keyword text."""
    key_ag = key_ag.lower().strip()
    # 1) Geo in keyword?
    for geo_phrase, our_ag in GEO_MAP.items():
        if geo_phrase in keyword_lower:
            return our_ag
    # 2) Exact key ad_group match
    if key_ag in KEY_AG_TO_OUR_AG:
        return KEY_AG_TO_OUR_AG[key_ag]
    # 3) Partial key ad_group match
    for k, our_ag in KEY_AG_TO_OUR_AG.items():
        if k in key_ag or key_ag in k:
            return our_ag
    # 4) Keyword-based fallback
    if "near me" in keyword_lower or "my area" in keyword_lower:
        return "Core Near Me"
    if "emergency" in keyword_lower or "same day" in keyword_lower or "urgent" in keyword_lower:
        return "Core Emergency"
    if "quote" in keyword_lower or "cost" in keyword_lower or "estimate" in keyword_lower or "price" in keyword_lower:
        return "Core Cost Quote"
    if "freezer" in keyword_lower:
        if "ice" in keyword_lower or "dispenser" in keyword_lower:
            return "Ice Maker Repair"
        if "not cool" in keyword_lower or "not cold" in keyword_lower or "not freez" in keyword_lower:
            return "Not Cooling Issues" if "cool" in keyword_lower else "Not Freezing Issues"
        return "Freezer Repair"
    if "ice" in keyword_lower and ("maker" in keyword_lower or "dispenser" in keyword_lower):
        return "Ice Maker Repair"
    if "leak" in keyword_lower or "water" in keyword_lower and ("not work" in keyword_lower or "dispenser" in keyword_lower):
        return "Water Leak Repair"
    if "door" in keyword_lower or "seal" in keyword_lower or "gasket" in keyword_lower or "handle" in keyword_lower:
        return "Door Seal Repair"
    if "wine" in keyword_lower:
        return "Wine Cooler Repair"
    if "technician" in keyword_lower or "repairman" in keyword_lower or "mechanic" in keyword_lower:
        return "Technician Request"
    if "noise" in keyword_lower or "compressor" in keyword_lower or "beep" in keyword_lower:
        return "Noise Compressor Fan"
    if "temperature" in keyword_lower or "warm" in keyword_lower or "too cold" in keyword_lower or "light not" in keyword_lower:
        return "Temperature Issues"
    if "not cool" in keyword_lower or "not cold" in keyword_lower:
        return "Not Cooling Issues"
    if "not freez" in keyword_lower:
        return "Not Freezing Issues"
    if "not work" in keyword_lower or "stopped" in keyword_lower or "broken" in keyword_lower:
        return "Not Working Issues"
    if "service" in keyword_lower and "repair" not in keyword_lower:
        return "Core Brand Service"
    return "Core Brand Repair"


def load_key_xlsx():
    import sys
    sys.path.insert(0, str(BASE / ".pydeps"))
    import openpyxl
    wb = openpyxl.load_workbook(KEY_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() for h in rows[0]]
    kw_i = next(i for i, h in enumerate(header) if "Keyword" in h)
    ag_i = next(i for i, h in enumerate(header) if "Ad Group" in h)
    seen = set()
    result = []
    for r in rows[1:]:
        kw = (r[kw_i] or "").strip()
        if not kw:
            continue
        k = normalize_kw(kw)
        if "sub zero" not in k and "subzero" not in k and "sub-zero" not in k:
            continue
        if k in seen:
            continue
        seen.add(k)
        ag = (r[ag_i] or "").strip()
        our_ag = map_key_ag_to_our_ag(ag, k)
        result.append((kw, our_ag))
    return result


def expand_to_5000(seed_list):
    """Expand seed (keyword, ad_group) list to 5000 unique keywords. Add variations."""
    # Normalize to canonical form for dedup (lower, single space)
    by_canonical = {}
    for kw, ag in seed_list:
        c = normalize_kw(kw)
        if c not in by_canonical:
            by_canonical[c] = (kw, ag)

    # Brand spelling variants for a phrase (sub zero / subzero / sub-zero)
    def add_variants(phrase, ag):
        out = []
        low = phrase.lower()
        for brand in ["sub zero", "subzero", "sub-zero"]:
            if "sub zero" in low:
                new = phrase.replace("sub zero", brand).replace("Sub Zero", brand.title())
            elif "subzero" in low:
                new = phrase.replace("subzero", brand).replace("Subzero", brand.title())
            elif "sub-zero" in low:
                new = phrase.replace("sub-zero", brand).replace("Sub-Zero", brand.title())
            else:
                new = phrase
            c = normalize_kw(new)
            if c not in by_canonical:
                out.append((new, c, ag))
        return out

    added = []
    for kw, ag in list(by_canonical.values()):
        for new_kw, c, a in add_variants(kw, ag):
            if c not in by_canonical:
                by_canonical[c] = (new_kw, a)
                added.append((new_kw, a))

    # Geo expansion: take high-intent base phrases and add city names
    base_phrases = []
    for kw, ag in by_canonical.values():
        k = kw.lower()
        if "repair" in k and "near me" in k and ag == "Core Near Me":
            base_phrases.append((kw, ag))
        if "repair" in k and "service" in k and ag in ("Core Brand Repair", "Core Near Me"):
            base_phrases.append((kw, ag))

    cities = list(GEO_MAP.keys())
    for phrase, _ in base_phrases[:80]:
        for city in cities:
            if city in phrase:
                continue
            # "sub zero refrigerator repair near me" -> "sub zero refrigerator repair san jose"
            new = phrase.replace(" near me", " " + city).replace(" my area", " " + city)
            c = normalize_kw(new)
            if c not in by_canonical:
                our_ag = GEO_MAP.get(city, "Core Brand Repair")
                by_canonical[c] = (new, our_ag)
                if len(by_canonical) >= 5000:
                    break
        if len(by_canonical) >= 5000:
            break

    # If still under 5000, add more intent/word order variants from seed
    seeds = list(by_canonical.values())
    templates = [
        ("sub zero {} repair", "Core Brand Repair"),
        ("sub zero {} repair near me", "Core Near Me"),
        ("sub zero {} fix", "Core Brand Repair"),
        ("sub zero {} service", "Core Brand Service"),
        ("{} sub zero repair", "Core Brand Repair"),
        ("repair sub zero {}", "Core Brand Repair"),
    ]
    products = ["refrigerator", "freezer", "appliance", "fridge", "wine cooler", "built in refrigerator"]
    for tpl, ag in templates:
        for p in products:
            kw = tpl.format(p)
            c = normalize_kw(kw)
            if c not in by_canonical:
                by_canonical[c] = (kw, ag)
            if len(by_canonical) >= 5000:
                break
        if len(by_canonical) >= 5000:
            break

    # Emergency / same day
    for kw, ag in list(by_canonical.values())[:200]:
        if "repair" not in kw.lower():
            continue
        for extra in [" emergency", " same day", " urgent", " today"]:
            new = kw.rstrip(".") + extra
            c = normalize_kw(new)
            if c not in by_canonical:
                by_canonical[c] = (new, "Core Emergency")
            if len(by_canonical) >= 5000:
                break
        if len(by_canonical) >= 5000:
            break

    # Not cooling / not freezing / not working variants
    for kw, ag in list(by_canonical.values()):
        if len(by_canonical) >= 5000:
            break
        k = kw.lower()
        if "not cooling" in k or "not cold" in k:
            for v in ["sub zero refrigerator not cooling", "sub zero fridge not cooling", "subzero refrigerator not cold"]:
                c = normalize_kw(v)
                if c not in by_canonical:
                    by_canonical[c] = (v, "Not Cooling Issues")
        if "not freezing" in k or "not making ice" in k:
            for v in ["sub zero freezer not freezing", "sub zero ice maker not making ice"]:
                c = normalize_kw(v)
                if c not in by_canonical:
                    by_canonical[c] = (v, "Not Freezing Issues" if "freez" in v else "Ice Maker Repair")
        if "ice maker" in k:
            for v in ["sub zero ice maker not working", "sub zero refrigerator ice maker repair"]:
                c = normalize_kw(v)
                if c not in by_canonical:
                    by_canonical[c] = (v, "Ice Maker Repair")

    # Pad to 5000: add "sub zero X" / "subzero X" variants and city suffixes
    cities = list(GEO_MAP.keys())
    seeds_ordered = list(by_canonical.values())
    for kw, ag in seeds_ordered:
        if len(by_canonical) >= 5000:
            break
        k = kw.lower()
        # City suffix variants (for Core Near Me -> Geo)
        for city in cities:
            if len(by_canonical) >= 5000:
                break
            if city in k:
                continue
            new = f"{kw} {city}"
            c = normalize_kw(new)
            if c not in by_canonical:
                by_canonical[c] = (new, GEO_MAP.get(city, ag))
    # Product + intent matrix
    for product in ["refrigerator", "freezer", "fridge", "wine cooler"]:
        if len(by_canonical) >= 5000:
            break
        for intent in ["repair", "repair near me", "service", "fix", "technician", "repair service"]:
            for brand in ["sub zero", "subzero", "sub-zero"]:
                kw = f"{brand} {product} {intent}"
                c = normalize_kw(kw)
                if c not in by_canonical:
                    by_canonical[c] = (kw, "Core Near Me" if "near me" in intent else "Core Brand Repair")
                if len(by_canonical) >= 5000:
                    break

    out = []
    for c, (kw, ag) in list(by_canonical.items())[:5000]:
        out.append((kw, ag))
    return out[:5000]


def get_cpc_for_ad_group(ad_group: str) -> str:
    for ag, cpc in AD_GROUPS_CORE:
        if ag == ad_group:
            return f"{cpc:.2f}"
    return "36.00"


def main():
    PACKAGE.mkdir(parents=True, exist_ok=True)
    seed = load_key_xlsx()
    print(f"Loaded {len(seed)} unique branded keywords from key.xlsx")
    keywords_with_ag = expand_to_5000(seed)
    # Dedupe and cap at 5000
    seen_k = set()
    unique_5000 = []
    for kw, ag in keywords_with_ag:
        c = normalize_kw(kw)
        if c in seen_k:
            continue
        seen_k.add(c)
        unique_5000.append((kw.strip(), ag))
        if len(unique_5000) == 5000:
            break
    # Pad to exactly 5000 with generated variants
    brands = ["sub zero", "subzero", "sub-zero"]
    products = ["refrigerator", "freezer", "fridge", "wine cooler", "appliance"]
    intents = ["repair", "repair near me", "service", "fix", "repair service", "technician", "repairman"]
    idx = 0
    while len(unique_5000) < 5000:
        b, p, i = brands[idx % 3], products[(idx // 3) % 5], intents[(idx // 15) % 7]
        kw = f"{b} {p} {i}"
        c = normalize_kw(kw)
        if c not in seen_k:
            seen_k.add(c)
            ag = "Core Near Me" if "near me" in i else "Core Brand Repair"
            unique_5000.append((kw, ag))
        idx += 1
        if idx > 50000:
            break
    unique_5000 = unique_5000[:5000]

    # --- 01_campaign.csv ---
    with open(PACKAGE / "01_campaign.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Campaign", "Campaign type", "Campaign daily budget", "Bid strategy type", "Networks", "Languages", "Start date", "Campaign status", "Comment"])
        w.writerow([
            CAMPAIGN_NAME, "Search", "600.00", "Manual CPC",
            "Google Search;Search Partners", "en", "2026-02-23", "Enabled",
            "Exact + Phrase only. 5000 unique keywords x2 = 10000 rows. No broad. Bay Area Sub-Zero repair leads."
        ])

    # --- 02_ad_groups.csv ---
    with open(PACKAGE / "02_ad_groups.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Campaign", "Ad Group", "Max CPC", "Campaign Status", "Ad Group Status", "Comment"])
        for ag, cpc in AD_GROUPS_CORE:
            w.writerow([CAMPAIGN_NAME, ag, f"{cpc:.2f}", "Enabled", "Enabled", f"Intent/Geo: {ag}"])

    # --- 03_keywords: 10000 rows (5000 Exact + 5000 Phrase) ---
    with open(PACKAGE / "03_keywords_10000_master.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Campaign", "Ad Group", "Keyword", "Criterion Type", "Max CPC", "Final URL", "Campaign Status", "Ad Group Status", "Status", "Comment"])
        for kw, ag in unique_5000:
            cpc = get_cpc_for_ad_group(ag)
            for criterion_type in ["Exact", "Phrase"]:
                w.writerow([
                    CAMPAIGN_NAME, ag, kw, criterion_type, cpc, FINAL_URL,
                    "Enabled", "Enabled", "Enabled", f"{ag} | {criterion_type}"
                ])

    # --- 03a_keywords_core_exact_phrase.csv: same as 03 for Editor (single file import) ---
    with open(PACKAGE / "03a_keywords_core_exact_phrase.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Campaign", "Ad Group", "Keyword", "Criterion Type", "Max CPC", "Final URL", "Campaign Status", "Ad Group Status", "Status", "Comment"])
        for kw, ag in unique_5000:
            cpc = get_cpc_for_ad_group(ag)
            for criterion_type in ["Exact", "Phrase"]:
                w.writerow([
                    CAMPAIGN_NAME, ag, kw, criterion_type, cpc, FINAL_URL,
                    "Enabled", "Enabled", "Enabled", f"{ag} | {criterion_type}"
                ])

    print(f"Written 01_campaign.csv (1 campaign), 02_ad_groups.csv ({len(AD_GROUPS_CORE)} groups)")
    print(f"Written 03_keywords_10000_master.csv and 03a_keywords_core_exact_phrase.csv: {len(unique_5000)} unique -> {len(unique_5000)*2} rows")
    return unique_5000


if __name__ == "__main__":
    main()
